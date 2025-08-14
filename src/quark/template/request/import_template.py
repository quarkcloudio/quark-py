from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from quark import Request, Response


class ImportTemplateRequest:

    # 请求对象
    request: Request = None

    # 列表页字段
    fields: list = None

    def __init__(
        self,
        request: Request,
        fields: list,
    ):
        self.request = request
        self.fields = fields

    def handle(self):
        """
        处理导出导入模板逻辑
        """
        export_titles = [
            self.get_field_label_with_remark(field) for field in self.fields
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 写入表头
        for col_idx, title in enumerate(export_titles, start=1):
            ws[f"{get_column_letter(col_idx)}1"] = title

        # 返回 Excel 文件流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"data_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/octet-stream",
        }

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    def get_field_label_with_remark(self, field):
        """
        获取字段 label + 提示信息
        """
        label = field.get("label", "")
        remark = self.get_field_remark(field)
        if remark:
            return f"{label}（{remark}）"
        return label

    def get_field_remark(self, field):
        """
        根据字段组件类型和规则生成备注信息
        """
        component = field.get("component")
        rules = field.get("rules", [])
        mode = field.get("mode")
        options = field.get("options", "")

        remarks = []

        if component == "inputNumberField":
            remarks.append("数字格式")
        elif component == "textField":
            pass  # 无特殊说明
        elif component == "selectField":
            if mode == "multiple":
                remarks.append(f"可多选：{options}；多值请用“,”分割")
            else:
                remarks.append(f"可选项：{options}")
        elif component == "cascaderField":
            remarks.append("级联格式，例如：省，市，县")
        elif component == "checkboxField":
            remarks.append(f"可多选项：{options}；多值请用“,”分割")
        elif component == "radioField":
            remarks.append(f"可选项：{options}")
        elif component == "switchField":
            remarks.append(f"可选项：{options}")
        elif component == "dateField":
            remarks.append("日期格式，例如：1987-02-15")
        elif component == "datetimeField":
            remarks.append("日期时间格式，例如：1987-02-15 20:00:00")

        rule_messages = self.get_rule_messages(rules)
        if rule_messages:
            remarks.append(f"条件：{rule_messages}")

        return "，".join(remarks)

    def get_rule_messages(self, rules):
        """
        解析规则并返回中文描述
        """
        messages = []
        for rule in rules:
            if isinstance(rule, dict):
                rule_type = rule.get("type")
                value = rule.get("value")
                if rule_type == "required":
                    messages.append("必填")
                elif rule_type == "min":
                    messages.append(f"大于{value}个字符")
                elif rule_type == "max":
                    messages.append(f"小于{value}个字符")
                elif rule_type == "email":
                    messages.append("必须为邮箱格式")
                elif rule_type == "numeric":
                    messages.append("必须为数字格式")
                elif rule_type == "url":
                    messages.append("必须为链接格式")
                elif rule_type == "integer":
                    messages.append("必须为整数格式")
                elif rule_type == "date":
                    messages.append("必须为日期格式")
                elif rule_type == "boolean":
                    messages.append("必须为布尔格式")
                elif rule_type == "unique":
                    messages.append("不可重复")
        return "，".join(messages)
