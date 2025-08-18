import json
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tortoise.queryset import QuerySet

from quark import Request, Response

from ...services.attachment import AttachmentService
from ..performs_queries import PerformsQueries
from ..resolves_actions import ResolvesActions


class ExportRequest:

    # 请求对象
    request: Request = None

    # 查询对象
    query: QuerySet = None

    # 列表页字段
    fields: list = None

    # 搜索项
    searches: list = None

    # 全局数据排序规则
    query_order: str = None

    # 列表页面的排序规则
    export_query_order: str = None

    def __init__(
        self,
        request: Request,
        query: QuerySet,
        query_order: str,
        export_query_order: str,
        fields: list,
        searches: list,
    ):
        self.request = request
        self.query = query
        self.query_order = query_order
        self.export_query_order = export_query_order
        self.fields = fields
        self.searches = searches

    def handle(self):
        """
        处理导出逻辑
        """
        data = self.query_data()
        fields = self.fields

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 写入表头
        for col_idx, field in enumerate(fields, start=1):
            label = field.get("label", "")
            ws[f"{get_column_letter(col_idx)}1"] = label

        # 写入数据行
        for row_idx, item in enumerate(data, start=2):
            for col_idx, field in enumerate(fields, start=1):
                name = field.get("name")
                component = field.get("component")

                value = None
                if component == "inputNumberField" or component == "textField":
                    value = item.get(name)
                elif component in ["selectField", "checkboxField", "radioField"]:
                    option_label = field.get("getOptionLabel")(item.get(name))
                    value = option_label
                elif component == "switchField":
                    option_label = field.get("getOptionLabel")(item.get(name))
                    value = option_label
                else:
                    value = item.get(name)

                # 特殊格式处理
                if component in ["datetimeField", "dateField"]:
                    format_str = field.get("format", "")
                    format_str = (
                        format_str.replace("YYYY", "%Y")
                        .replace("MM", "%m")
                        .replace("DD", "%d")
                        .replace("HH", "%H")
                        .replace("mm", "%M")
                        .replace("ss", "%S")
                    )
                    if isinstance(value, datetime):
                        value = value.strftime(format_str)

                ws[f"{get_column_letter(col_idx)}{row_idx}"] = value

        # 返回 Excel 文件流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"data_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/octet-stream",
        }

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    async def query_data(self):
        """
        查询数据
        """

        # 获取列过滤条件
        column_filters = self.column_filters()

        # 获取排序规则
        orderings = self.orderings()

        query = PerformsQueries(
            request=self.request,
            query=self.query,
        ).build_export_query(self.searches, column_filters)

        query = PerformsQueries(
            request=self.request,
            query_order=self.query_order,
            export_query_order=self.export_query_order,
        ).apply_index_orderings(query, orderings)

        result = query.all()

        return await self.performs_list(result)

    async def performs_list(self, items):
        """
        处理列表字段
        """
        result = []
        index_fields = self.fields
        for item in items:
            fields = {}

            for field in index_fields:
                component = field.component
                name = field.name

                if component == "actionField":
                    items_callback = field.callback
                    if items_callback:
                        action_items = await items_callback(item)
                    else:
                        action_items = field.items

                    rendered_actions = []
                    for action in action_items:
                        rendered_actions.append(
                            await ResolvesActions(self.request).build_action(action)
                        )

                    fields[name] = rendered_actions
                else:
                    callback = field.callback
                    if callback:
                        fields[name] = await callback(item)
                    else:
                        value = getattr(item, name, None)
                        if value is None:
                            continue

                        # JSON 字符串解析
                        if isinstance(value, str):
                            if value.startswith("[") or value.startswith("{"):
                                try:
                                    value = json.loads(value)
                                except:
                                    pass

                        # 图片字段处理
                        if component in ["imageField", "imagePickerField"]:
                            value = AttachmentService().get_image_url(value)

                        fields[name] = value

            result.append(fields)

        return result

    def column_filters(self):
        """
        获取列过滤条件
        """
        filter_str = self.request.query_params.get("filter")
        column_filters = {}
        try:
            column_filters = json.loads(filter_str)
        except:
            pass
        return column_filters

    def orderings(self):
        """
        获取排序规则
        """
        sorter_str = self.request.query_params.get("sorter")
        orderings = {}
        try:
            orderings = json.loads(sorter_str)
        except:
            pass
        return orderings
