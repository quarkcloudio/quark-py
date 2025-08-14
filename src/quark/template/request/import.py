import json
import os
import random
import string

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from tortoise.models import Model

from quark import Message, Request


class ImportRequest:

    # 请求对象
    request: Request = None

    # 查询对象
    query: Model = None

    # 列表页字段
    fields: list = None

    def __init__(
        self,
        request: Request,
        query: Model,
        fields: list,
    ):
        self.request = request
        self.query = query
        self.fields = fields

    def handle(self, index_route):
        """
        处理导入请求
        """
        body_bytes = self.request.body()
        data = json.loads(body_bytes)
        file_ids = data.get("fileId", [])

        if not file_ids or not isinstance(file_ids, list) or len(file_ids) == 0:
            return Message.error("参数错误")

        file_id_info = file_ids[0]
        file_id = file_id_info.get("id")
        if not file_id:
            return Message.error("参数错误")

        # 获取Excel数据
        import_data = self.get_excel_data(file_id)
        if not import_data:
            return Message.error("文件数据为空")

        # 表头与数据分离
        import_head = import_data[0]
        import_data = import_data[1:]

        # 导入前回调
        lists = template.before_importing(import_data)

        import_result = True
        total_num = len(lists)
        success_num = 0
        failed_num = 0
        failed_data = []

        fields = self.fields

        for item in lists:
            form_values = self.transform_form_values(fields, item)

            # 验证器
            validator = template.validator_for_import(form_values)
            if validator:
                import_result = False
                failed_num += 1
                padding_len = len(fields) - len(item)
                item += [None] * padding_len
                item.append(validator.error())
                failed_data.append(item)
                continue

            # BeforeSaving 回调
            try:
                submit_data, err = template.before_saving(form_values)
                if err:
                    raise Exception(err)
            except Exception as e:
                import_result = False
                failed_num += 1
                padding_len = len(fields) - len(item)
                item += [None] * padding_len
                item.append(str(e))
                failed_data.append(item)
                continue

            # 提交数据
            data_to_submit = self.get_submit_data(fields, submit_data)
            try:
                record = YourModel(**data_to_submit)
                db.session.add(record)
                db.session.commit()

                # AfterImported 回调
                last_id = record.id
                after_err = template.after_imported(last_id, data_to_submit)
                if after_err:
                    raise Exception(after_err)

                success_num += 1
            except Exception as e:
                db.session.rollback()
                import_result = False
                failed_num += 1
                padding_len = len(fields) - len(item)
                item += [None] * padding_len
                item.append(str(e))
                failed_data.append(item)

        # 返回失败数据
        if not import_result:
            static_path = app.config["STATIC_PATH"]
            storage_dir = os.path.join(static_path, "app", "storage", "failImports")
            file_name = self.generate_random_filename(40) + ".xlsx"
            file_url = f"//{request.host}/storage/failImports/{file_name}"

            if not os.path.exists(storage_dir):
                os.makedirs(storage_dir)

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # 写入表头
            import_head.append("错误信息")
            for col_idx, title in enumerate(import_head, start=1):
                ws[f"{generate_column_label(col_idx)}1"] = title

            # 写入失败数据
            for row_idx, row in enumerate(failed_data, start=2):
                for col_idx, cell_value in enumerate(row, start=1):
                    ws[f"{generate_column_label(col_idx)}{row_idx}"] = cell_value
                    if col_idx == len(row):
                        cell = ws[f"{generate_column_label(col_idx)}{row_idx}"]
                        cell.font = Font(color="FF0000")

            # 保存文件
            file_path = os.path.join(storage_dir, file_name)
            wb.save(file_path)

            response_html = (
                f"<div style='margin: 20px;'>"
                f"<p>导入总量: {total_num}</p>"
                f"<p>成功数量: {success_num}</p>"
                f"<p>失败数量: <span style='color:#ff4d4f'>{failed_num}</span> "
                f"<a href='{file_url}' target='_blank'>下载失败数据</a></p>"
                f"</div>"
            )

            return jsonify({"html": response_html}), 200

        return redirect(
            f"/layout/index?api={index_route.replace(':resource', request.view_args.get('resource'))}"
        )

    def get_excel_data(self, file_id):
        """
        获取附件服务中的 Excel 数据（需根据你的业务系统实现）
        """
        from your_app.models import Attachment

        attachment = Attachment.query.get(file_id)
        if not attachment:
            return []
        file_path = attachment.file_path
        if not os.path.exists(file_path):
            return []
        # 使用 openpyxl 或 pandas 读取 Excel 数据

        df = pd.read_excel(file_path, header=None)
        return df.values.tolist()

    def transform_form_values(self, fields, data):
        result = {}
        for idx, field in enumerate(fields):
            name = field.get("name")
            if idx < len(data):
                result[name] = data[idx]
        return result

    def get_submit_data(self, fields, submit_data):
        result = {}

        for field in fields:
            component = field.get("component")
            name = field.get("name")
            value = submit_data.get(name)

            if component == "inputNumberField":
                result[name] = value
            elif component == "textField":
                result[name] = str(value).strip("\n") if value else None
            elif component in ["selectField", "checkboxField", "radioField"]:
                label = value
                option_value = field.get("getOptionValue")(label)
                result[name] = option_value
            elif component == "switchField":
                label = value
                option_value = field.get("getOptionValue")(label)
                result[name] = option_value
            else:
                result[name] = value

            # 转换复杂类型为 JSON 字符串
            if isinstance(result[name], (list, dict)):
                result[name] = json.dumps(result[name])

        return result

    def generate_random_filename(self, length=40):
        chars = string.ascii_letters + string.digits
        return "".join(random.choice(chars) for _ in range(length))
